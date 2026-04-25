"""Sync the founder IVIS workbook into a runtime JSON catalog.

This is a one-way export utility:
  workbook (.xlsx) -> config/ivis_dictionary.json

The runtime product reads the generated JSON file rather than depending on
spreadsheet tooling in production. The workbook remains the founder-editable
source of truth for copy.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - script guard
    raise SystemExit(
        "openpyxl is required to sync the IVIS workbook. Install it first."
    ) from exc


RUNTIME_EVENT_MAP: dict[str, str] = {
    # Current runtime-detected scenarios supported by founder-approved copy.
    "stall": "ME-31",
    "emergency": "SO-27",
}


def _build_mode_metadata(legend_rows: list[tuple[Any, ...]]) -> dict[str, dict[str, str]]:
    mode_metadata: dict[str, dict[str, str]] = {}
    for row in legend_rows:
        label = row[0]
        if not isinstance(label, str):
            continue

        if label.startswith("MODE 1"):
            mode_metadata["mode_1"] = {
                "description": row[1] or "",
                "regulatory_principle": row[2] or "",
            }
        elif label.startswith("MODE 2"):
            mode_metadata["mode_2"] = {
                "description": row[1] or "",
                "regulatory_principle": row[2] or "",
            }
        elif label.startswith("MODE 3"):
            mode_metadata["mode_3"] = {
                "description": row[1] or "",
                "regulatory_principle": row[2] or "",
            }
    return mode_metadata


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def export_dictionary(source: Path, output: Path) -> dict[str, Any]:
    workbook = load_workbook(source, data_only=True)
    rules_sheet = workbook["IVIS Dictionary v2"]
    legend_sheet = workbook["Legend & Regulatory Notes"]

    legend_rows = list(legend_sheet.iter_rows(values_only=True))
    mode_metadata = _build_mode_metadata(legend_rows)

    rules: list[dict[str, Any]] = []
    for row in rules_sheet.iter_rows(min_row=3, values_only=True):
        rule_id = _clean_text(row[2])
        scenario = _clean_text(row[3])
        if not rule_id or not scenario:
            continue

        runtime_events = [
            event_type
            for event_type, mapped_rule_id in RUNTIME_EVENT_MAP.items()
            if mapped_rule_id == rule_id
        ]

        rules.append(
            {
                "rule_id": rule_id,
                "category": _clean_text(row[1]),
                "scenario": scenario,
                "what_needs_to_be_done": _clean_text(row[4]),
                "driver_state": _clean_text(row[5]),
                "typical_human_reaction": _clean_text(row[6]),
                "ideal_human_response": _clean_text(row[7]),
                "reference_response": _clean_text(row[8]),
                "mode_1": _clean_text(row[9]),
                "mode_2": _clean_text(row[10]),
                "mode_3": _clean_text(row[11]),
                "runtime_event_types": runtime_events,
            }
        )

    payload = {
        "metadata": {
            "source_workbook": source.name,
            "source_sheet": "IVIS Dictionary v2",
            "legend_sheet": "Legend & Regulatory Notes",
            "total_rules": len(rules),
            "mode_metadata": mode_metadata,
        },
        "runtime_event_map": {
            event_type: {"rule_id": rule_id}
            for event_type, rule_id in RUNTIME_EVENT_MAP.items()
        },
        "rules": rules,
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Path to IVIS_Dictionary_v2.xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("config/ivis_dictionary.json"),
        help="Output JSON catalog path",
    )
    args = parser.parse_args()

    payload = export_dictionary(args.source, args.output)
    print(f"Exported {payload['metadata']['total_rules']} IVIS rules to {args.output}")
    return 0


if __name__ == "__main__":  # pragma: no cover - script entry point
    raise SystemExit(main())
